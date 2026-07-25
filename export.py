"""Stage 3 — styled Excel workbook export.

Tabs: Summary · Methodology · Confidence · Representativeness · Analysis Summary ·
Market Intelligence (Cited) · Run Log · one data tab per channel (with analysis columns
joined). The tool version is stamped into the workbook so any client-facing report is
traceable to the exact build that produced it. Optional published-date filter.

Every export documents its own methodology AND its gaps — the Methodology, Confidence,
and Representativeness tabs are not optional decoration; they are the honesty contract.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import analytics
import storage
from scrapers import CHANNEL_INFO
from settings import settings
from version import __version__

ANALYSIS_CAVEAT = (
    "Analysis tags (sentiment, drivers, themes, translations) are model-generated. "
    "Expect ~85–95% agreement with a human coder. Spot-check before quoting any tag or "
    "verbatim in a client deliverable."
)

REPRESENTATIVENESS_CAVEAT = (
    "All automated digital sources skew toward urban, online-active, and literate-in-"
    "covered-languages populations. Findings describe the digitally-expressed segment of "
    "the market, not the whole market. Weight accordingly; triangulate with the cited "
    "layer and offline research."
)


def _styles():
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1F3864")
    wrap = Alignment(vertical="top", wrap_text=True)
    flag_fill = PatternFill("solid", fgColor="F8CBAD")  # low-confidence flag
    return {"header_fill": header_fill, "header_font": header_font, "title_font": title_font,
            "wrap": wrap, "flag_fill": flag_fill}


def _write_header(ws, headers, row=1, styles=None):
    styles = styles or _styles()
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = styles["header_fill"]
        c.font = styles["header_font"]
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, max_width=60):
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        length = 10
        letter = None
        for cell in col:
            if letter is None:
                letter = get_column_letter(cell.column)
            try:
                length = max(length, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        if letter:
            ws.column_dimensions[letter].width = min(length + 2, max_width)


DATA_COLUMNS = [
    ("id", "id"), ("source", "source"), ("title", "title"), ("text", "text"),
    ("link", "link"), ("published", "published"), ("run_id", "run_id"),
    ("sentiment", "sentiment"), ("sentiment_score", "sentiment_score"), ("language", "language"),
    ("summary_en", "summary_en"), ("rating_signal", "rating_signal"),
    ("purchase_driver", "purchase_driver"), ("usage_occasion", "usage_occasion"),
    ("trend_category", "trend_category"), ("brand_focus", "brand_focus"),
    ("promo_mentioned", "promo_mentioned"), ("emotion", "emotion"),
]


def build_workbook(project_id: int, published_after: Optional[str] = None,
                   published_before: Optional[str] = None,
                   out_path: Optional[str] = None) -> str:
    from openpyxl import Workbook

    project = storage.get_project(project_id)
    if not project:
        raise ValueError(f"Project {project_id} not found")
    cfg = project["config"]
    styles = _styles()

    wb = Workbook()
    _summary_tab(wb.active, project, cfg, styles, published_after, published_before)
    _methodology_tab(wb.create_sheet("Methodology"), project_id, cfg, styles)
    _confidence_tab(wb.create_sheet("Confidence"), project_id, styles)
    _representativeness_tab(wb.create_sheet("Representativeness"), cfg, styles)
    _analysis_summary_tab(wb.create_sheet("Analysis Summary"), project_id, styles)
    _market_intel_tab(wb.create_sheet("Market Intelligence (Cited)"), project_id, styles)
    _run_log_tab(wb.create_sheet("Run Log"), project_id, styles)

    # Combined master tab: EVERY item across all channels, with analysis columns joined.
    all_rows = _filter_dates(storage.items_with_analysis(project_id), published_after, published_before)
    if all_rows:
        _channel_data_tab(wb.create_sheet("All Items"), all_rows, styles)

    # One data tab per channel that actually has items.
    counts = storage.count_items_by_source(project_id)
    for channel in sorted(counts):
        rows = storage.items_with_analysis(project_id, source=channel)
        rows = _filter_dates(rows, published_after, published_before)
        if not rows:
            continue
        _channel_data_tab(wb.create_sheet(_safe_sheet_name(channel)), rows, styles)

    settings.ensure_dirs()
    if out_path is None:
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        safe = "".join(c for c in project["name"] if c.isalnum() or c in "-_") or "project"
        out_path = str(settings.exports_dir / f"MarketLens_{safe}_{ts}.xlsx")
    wb.save(out_path)
    storage.audit("export", f"Excel export -> {Path(out_path).name}", project_id=project_id)
    return out_path


def _filter_dates(rows, after, before):
    if not after and not before:
        return rows
    out = []
    for r in rows:
        pub = (r.get("published") or "")[:10]
        if after and pub and pub < after:
            continue
        if before and pub and pub > before:
            continue
        out.append(r)
    return out


def _safe_sheet_name(name: str) -> str:
    bad = set(r'[]:*?/\\')
    cleaned = "".join(c for c in name if c not in bad)
    return cleaned[:31] or "sheet"


# --------------------------------------------------------------------------- #
# Individual tabs
# --------------------------------------------------------------------------- #
def _summary_tab(ws, project, cfg, styles, after, before):
    ws.title = "Summary"
    ws["A1"] = "MarketLens — Study Summary"
    ws["A1"].font = styles["title_font"]
    market = cfg.get("market", {})
    product = cfg.get("product", {})
    counts = storage.count_items_by_source(project["id"])
    dash = analytics.dashboard(project["id"])

    rows = [
        ("Project", project["name"]),
        ("Tool version", f"MarketLens v{__version__}"),
        ("Generated (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("Date filter", f"{after or '—'} to {before or '—'}"),
        ("", ""),
        ("Brand", product.get("brand", "")),
        ("Parent company", product.get("parent_company", "")),
        ("Category", f"{product.get('category', '')} ({product.get('category_type', '')})"),
        ("Market", market.get("country", "")),
        ("Languages", ", ".join(market.get("languages", []))),
        ("Competitors", ", ".join(cfg.get("competitors", []))),
        ("", ""),
        ("Total items collected", sum(counts.values())),
        ("Items analyzed", dash["total_analyzed"]),
        ("Items awaiting analysis", dash["unanalyzed"]),
        ("Overall net sentiment", dash["overall_net_score"]),
        ("", ""),
    ]
    r = 3
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = styles["header_font"] if k and not v == "" else None
        ws.cell(row=r, column=2, value=v)
        r += 1
    ws.cell(row=r, column=1, value="Items by channel").font = styles["title_font"]
    r += 1
    _write_header(ws, ["Channel", "Items"], row=r, styles=styles)
    r += 1
    for ch in sorted(counts):
        ws.cell(row=r, column=1, value=ch)
        ws.cell(row=r, column=2, value=counts[ch])
        r += 1
    _autosize(ws)


def _methodology_tab(ws, project_id, cfg, styles):
    ws["A1"] = "Methodology & Limitations"
    ws["A1"].font = styles["title_font"]
    _write_header(ws, ["Channel", "Tier", "Method", "Limitations"], row=3, styles=styles)
    r = 4
    counts = storage.count_items_by_source(project_id)
    # Show every channel that ran, plus any Tier-1 channel for completeness.
    for ch in sorted(set(list(CHANNEL_INFO.keys()))):
        info = CHANNEL_INFO.get(ch, {})
        used = counts.get(ch, 0)
        ws.cell(row=r, column=1, value=f"{info.get('name', ch)}"
                                      + (f"  ({used} items)" if used else "  (not run)"))
        ws.cell(row=r, column=2, value=info.get("tier", ""))
        ws.cell(row=r, column=3, value=info.get("method", "")).alignment = styles["wrap"]
        ws.cell(row=r, column=4, value=info.get("limitation", "")).alignment = styles["wrap"]
        r += 2

    r += 1
    ws.cell(row=r, column=1, value="Tier-2 (Assisted-Manual) platforms").font = styles["title_font"]
    r += 1
    for p in cfg.get("source_plan", {}).get("manual_intelligence_platforms", []):
        ws.cell(row=r, column=1, value=p.get("name"))
        ws.cell(row=r, column=3, value=p.get("note", "")).alignment = styles["wrap"]
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Tier-3 — NOT COVERED (documented gaps)").font = styles["title_font"]
    r += 1
    _write_header(ws, ["Platform", "", "Reason not covered", ""], row=r, styles=styles)
    r += 1
    for gap in cfg.get("source_plan", {}).get("tier3_gaps", []):
        ws.cell(row=r, column=1, value=gap.get("platform"))
        ws.cell(row=r, column=3, value=gap.get("reason")).alignment = styles["wrap"]
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Analysis caveat").font = styles["title_font"]
    ws.cell(row=r + 1, column=1, value=ANALYSIS_CAVEAT).alignment = styles["wrap"]
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["C"].width = 70


def _confidence_tab(ws, project_id, styles):
    ws["A1"] = "Confidence — sample sizes & low-confidence flags"
    ws["A1"].font = styles["title_font"]
    ws["A2"] = ("Any channel or headline stat resting on < 100 items or a single segment is "
                "auto-flagged 'emerging / low-confidence'. Do not headline a flagged number.")
    ws["A2"].alignment = styles["wrap"]
    _write_header(ws, ["Scope", "n (items)", "Net sentiment", "Flag"], row=4, styles=styles)
    r = 5
    for ch in analytics.sentiment_by_channel(project_id):
        ws.cell(row=r, column=1, value=f"Channel: {ch['channel']}")
        ws.cell(row=r, column=2, value=ch["n"])
        ws.cell(row=r, column=3, value=ch["net_score"])
        flag = "LOW-CONFIDENCE (emerging)" if ch["low_confidence"] else "ok"
        cell = ws.cell(row=r, column=4, value=flag)
        if ch["low_confidence"]:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = styles["flag_fill"]
        r += 1

    drivers = analytics.top_purchase_drivers(project_id)
    ws.cell(row=r, column=1, value="Aggregate: purchase drivers")
    ws.cell(row=r, column=2, value=drivers["n"])
    ws.cell(row=r, column=4, value="LOW-CONFIDENCE" if drivers["low_confidence"] else "ok")
    if drivers["low_confidence"]:
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = styles["flag_fill"]
    r += 1
    for bvc in analytics.brand_vs_competitor_sentiment(project_id):
        ws.cell(row=r, column=1, value=f"Brand focus: {bvc['brand_focus']}")
        ws.cell(row=r, column=2, value=bvc["n"])
        ws.cell(row=r, column=3, value=bvc["net_score"])
        ws.cell(row=r, column=4, value="LOW-CONFIDENCE" if bvc["low_confidence"] else "ok")
        if bvc["low_confidence"]:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = styles["flag_fill"]
        r += 1
    _autosize(ws)


def _representativeness_tab(ws, cfg, styles):
    ws["A1"] = "Representativeness"
    ws["A1"].font = styles["title_font"]
    ws["A3"] = "Standing caveat"
    ws["A3"].font = styles["header_font"]
    ws["A4"] = REPRESENTATIVENESS_CAVEAT
    ws["A4"].alignment = styles["wrap"]
    ws.merge_cells("A4:F8")

    ws["A10"] = "Per-segment skew notes"
    ws["A10"].font = styles["header_font"]
    notes = [
        ("News / GDELT", "Reflects what outlets choose to publish; skews toward newsworthy events."),
        ("Reddit / Forums", "Skews younger, male, English-speaking, opinionated early adopters."),
        ("E-commerce reviews", "Skews toward buyers with strong (often extreme) opinions; snapshot only."),
        ("YouTube", "Comment culture skews toward engaged fans/critics, not median consumers."),
        ("Google Trends", "Relative search interest, not population sentiment; no demographics."),
        ("Manual / Cited layer", "Analyst-selected; coverage depends on analyst diligence."),
    ]
    r = 11
    _write_header(ws, ["Segment", "Skew note"], row=r, styles=styles)
    r += 1
    for seg, note in notes:
        ws.cell(row=r, column=1, value=seg)
        ws.cell(row=r, column=2, value=note).alignment = styles["wrap"]
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90


def _analysis_summary_tab(ws, project_id, styles):
    ws["A1"] = "Analysis Summary"
    ws["A1"].font = styles["title_font"]
    r = 3
    ws.cell(row=r, column=1, value="Sentiment by channel (n-size shown)").font = styles["header_font"]
    r += 1
    _write_header(ws, ["Channel", "n", "positive", "negative", "neutral", "mixed", "net"], row=r, styles=styles)
    r += 1
    for ch in analytics.sentiment_by_channel(project_id):
        for i, v in enumerate([ch["channel"], ch["n"], ch["positive"], ch["negative"],
                               ch["neutral"], ch["mixed"], ch["net_score"]], start=1):
            ws.cell(row=r, column=i, value=v)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Top purchase drivers").font = styles["header_font"]
    r += 1
    drivers = analytics.top_purchase_drivers(project_id)
    _write_header(ws, [f"Driver (n={drivers['n']})", "count"], row=r, styles=styles)
    r += 1
    for d in drivers["drivers"]:
        ws.cell(row=r, column=1, value=d["driver"])
        ws.cell(row=r, column=2, value=d["count"])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Trend volume (n-size per theme)").font = styles["header_font"]
    r += 1
    _write_header(ws, ["Trend category", "n"], row=r, styles=styles)
    r += 1
    for s in analytics.trend_volume_over_time(project_id)["series"]:
        ws.cell(row=r, column=1, value=s["trend_category"])
        ws.cell(row=r, column=2, value=s["n"])
        r += 1
    _autosize(ws)


def _market_intel_tab(ws, project_id, styles):
    ws["A1"] = "Market Intelligence (Cited)"
    ws["A1"].font = styles["title_font"]
    ws["A2"] = "Human-entered desk research. Every row carries its full citation."
    _write_header(ws, ["Category", "Metric", "Value", "Source", "Source URL",
                       "Published", "Accessed", "Confidence", "Entered by", "Notes"], row=4, styles=styles)
    r = 5
    for e in storage.list_market_intel(project_id, entry_type="cited"):
        for i, key in enumerate(["category", "metric", "value", "source_name", "source_url",
                                 "publication_date", "accessed_date", "confidence", "entered_by",
                                 "notes"], start=1):
            ws.cell(row=r, column=i, value=e.get(key))
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Manual Intelligence — ad observations (Tier-2)").font = styles["title_font"]
    r += 1
    _write_header(ws, ["Platform", "Advertiser", "Creative theme", "Format", "First seen",
                       "Entered by", "Notes"], row=r, styles=styles)
    r += 1
    for e in storage.list_market_intel(project_id, entry_type="manual_ad"):
        extra = e.get("extra", {})
        vals = [e.get("source_name"), e.get("value"), extra.get("creative_theme"),
                extra.get("format"), extra.get("first_seen_date"), e.get("entered_by"), e.get("notes")]
        for i, v in enumerate(vals, start=1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    _autosize(ws)


def _run_log_tab(ws, project_id, styles):
    ws["A1"] = "Run Log — full audit trail"
    ws["A1"].font = styles["title_font"]
    _write_header(ws, ["Run ID", "Channel", "Status", "Started", "Finished", "Returned",
                       "New", "Duplicate", "Triggered by", "Errors"], row=3, styles=styles)
    r = 4
    for run in storage.list_runs(project_id, limit=1000):
        vals = [run["id"], run["channel"], run["status"], run["started_at"], run["finished_at"],
                run["rows_returned"], run["rows_new"], run["rows_duplicate"], run.get("triggered_by"),
                (run.get("errors_json") or "")[:500]]
        for i, v in enumerate(vals, start=1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    _autosize(ws)


def _channel_data_tab(ws, rows, styles):
    headers = [h for h, _ in DATA_COLUMNS]
    _write_header(ws, headers, row=1, styles=styles)
    r = 2
    for row in rows:
        for col, (_, key) in enumerate(DATA_COLUMNS, start=1):
            val = row.get(key)
            if key == "text" and val:
                val = str(val)[:2000]
            ws.cell(row=r, column=col, value=val)
        r += 1
    _autosize(ws, max_width=50)
