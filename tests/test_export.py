"""Export workbook contains all required tabs + version stamp; report scaffold; citations."""
import analysis
import export
import market_intel
import report
import storage
from version import __version__

REQUIRED_TABS = ["Summary", "Methodology", "Confidence", "Representativeness",
                 "Analysis Summary", "Market Intelligence (Cited)", "Run Log"]


def _cfg():
    import config
    return config.run_wizard({
        "market": {"country": "Singapore", "languages": ["en"]},
        "product": {"brand": "Acme Cola", "category": "cola", "category_type": "fmcg_food"},
        "competitors": ["Fizzly"],
        "keywords": {"trend_terms": ["sugar-free"]},
    })


def _seed_and_analyze(pid):
    r = storage.start_run(pid, "news", {})
    storage.save_items(pid, r, "news", [
        {"title": "Acme Cola review", "text": "tasty and cheap", "link": "http://a", "published": "2024-03-01"},
        {"title": "Fizzly launch", "text": "new competitor drink", "link": "http://b", "published": "2024-03-02"},
    ])

    def call(prompt, model):
        import json
        n = prompt.count("] source=")
        return json.dumps([{
            "sentiment": "positive", "sentiment_score": 0.6, "language": "en",
            "summary_en": "positive review", "purchase_driver": "price",
            "trend_category": "sugar-free", "brand_focus": "target brand",
            "promo_mentioned": False, "emotion": "joy",
        } for _ in range(n)])

    analysis.analyze_all(pid, call_fn=call, model="test-model")


def test_workbook_has_all_required_tabs_and_version(fresh_db, tmp_path):
    from openpyxl import load_workbook

    pid = storage.create_project("Acme Study", _cfg())
    _seed_and_analyze(pid)
    market_intel.add_cited_entry(pid, {
        "category": "Market size", "metric": "SG cola market", "value": "S$500m",
        "source_name": "Statista", "source_url": "http://statista.example",
        "publication_date": "2024-01-01", "accessed_date": "2024-06-01", "confidence": "medium",
    }, entered_by="alice")

    path = tmp_path / "out.xlsx"
    export.build_workbook(pid, out_path=str(path))
    wb = load_workbook(str(path))

    for tab in REQUIRED_TABS:
        assert tab in wb.sheetnames, f"missing tab: {tab}"
    # One data tab per channel that has items.
    assert "news" in wb.sheetnames
    # Combined master tab with the full column set.
    assert "All Items" in wb.sheetnames
    headers = [c.value for c in wb["All Items"][1]]
    assert headers == ["id", "source", "title", "text", "link", "published", "run_id",
                       "story_group_size", "sentiment", "sentiment_score", "language",
                       "summary_en", "rating_signal", "purchase_driver", "usage_occasion",
                       "trend_category", "brand_focus", "promo_mentioned", "emotion"]
    # Version stamped into the Summary tab.
    summary_vals = [c.value for row in wb["Summary"].iter_rows() for c in row if c.value]
    assert any(f"v{__version__}" in str(v) for v in summary_vals)
    # Cited citation present.
    cited_vals = [c.value for row in wb["Market Intelligence (Cited)"].iter_rows() for c in row]
    assert "http://statista.example" in cited_vals


def test_export_surfaces_syndication_not_just_raw_count(fresh_db, tmp_path):
    from openpyxl import load_workbook

    pid = storage.create_project("Syndication Test", _cfg())
    r = storage.start_run(pid, "news", {})
    storage.save_items(pid, r, "news", [
        {"title": "Maggi price rises in KL - NST", "text": "x", "link": "http://a", "published": "2026-03-01"},
        {"title": "Maggi price rises in KL - Star", "text": "x", "link": "http://b", "published": "2026-03-01"},
        {"title": "Unrelated Maggi story", "text": "y", "link": "http://c", "published": "2026-03-20"},
    ])

    def call(prompt, model):
        import json
        n = prompt.count("] source=")
        return json.dumps([{"sentiment": "positive", "sentiment_score": 0.5, "language": "en",
                           "summary_en": "s", "purchase_driver": "price", "trend_category": "x",
                           "brand_focus": "target brand", "promo_mentioned": False,
                           "emotion": "joy"} for _ in range(n)])

    analysis.analyze_all(pid, call_fn=call, model="test-model")
    path = tmp_path / "syn.xlsx"
    export.build_workbook(pid, out_path=str(path))
    wb = load_workbook(str(path))

    # Summary tab states unique-stories distinctly from raw item count.
    summary_pairs = [(r[0].value, r[1].value) for r in wb["Summary"].iter_rows() if r[0].value]
    d = dict(summary_pairs)
    assert d.get("Total items collected") == 3
    assert d.get("Unique stories (syndication-adjusted)") == 2

    # Data tab carries a story_group_size column reflecting the reprint.
    ws = wb["All Items"]
    headers = [c.value for c in ws[1]]
    idx = headers.index("story_group_size")
    sizes = sorted(row[idx] for row in ws.iter_rows(min_row=2, values_only=True))
    assert sizes == [1, 2, 2]  # two reprints (size 2 each) + one standalone (size 1)


def test_cited_entry_requires_full_citation(fresh_db):
    pid = storage.create_project("P", _cfg())
    try:
        market_intel.add_cited_entry(pid, {"category": "Market size", "value": "x"})
        assert False, "should have raised"
    except ValueError as e:
        msg = str(e)
        assert "source_name" in msg and "source_url" in msg and "accessed_date" in msg


def test_report_draft_has_five_pillars_and_markers(fresh_db):
    pid = storage.create_project("P", _cfg())
    _seed_and_analyze(pid)
    md = report.draft_report(pid)
    for pillar in ["Market Overview", "Consumer Intelligence", "Key Trends",
                   "Product Innovation", "Partnership Opportunities"]:
        assert pillar in md
    assert "[ANALYST INPUT]" in md
    # Trend sub-section generated from configured trend term.
    assert "sugar-free" in md
    # n-sizes present.
    assert "n=" in md


def test_report_downloads_md_and_docx(fresh_db, tmp_path):
    pid = storage.create_project("P", _cfg())
    _seed_and_analyze(pid)

    md_path = report.save_markdown(pid, out_path=str(tmp_path / "r.md"))
    text = open(md_path, encoding="utf-8").read()
    assert "Market Overview" in text and "Consumer Intelligence" in text

    docx_path = report.save_docx(pid, out_path=str(tmp_path / "r.docx"))
    from docx import Document
    doc = Document(docx_path)
    headings = [p.text for p in doc.paragraphs if p.style.name.startswith("Heading")]
    assert any("Consumer Intelligence" in h for h in headings)
    assert any("Key Trends" in h for h in headings)


def test_manual_intelligence_plan_deep_links(fresh_db):
    cfg = _cfg()
    plan = market_intel.manual_intelligence_plan(cfg)
    meta = next(p for p in plan if p["key"] == "meta_ad_library")
    # Deep links built for brand + competitor.
    names = [dl["name"] for dl in meta["deep_links"]]
    assert "Acme Cola" in names and "Fizzly" in names
    assert all("facebook.com/ads/library" in dl["url"] for dl in meta["deep_links"])
